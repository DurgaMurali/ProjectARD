import openpyxl
import pandas

from bbc.accounts_receivable import AccountsReceivable, MarginableAR

def read_accounts_receivable(file_name) -> AccountsReceivable:
        
           # Read insured AR
    # To open the workbook
    # workbook object is created
    wb_obj = openpyxl.load_workbook(file_name)
    sheet_obj = wb_obj.get_sheet_by_name("Insured AR")
    max_col = sheet_obj.max_column
    max_row = sheet_obj.max_row

    # Loop will print all columns name
    # for i in range(1, max_col + 1): 
    #     for j in range(1, max_row + 1):
    #         cell_obj = sheet_obj.cell(row=j, column=i)
    #         if cell_obj != None
    #          header_row = j
    #         break
    #         else
    #          return
        
    file_AR = pandas.read_excel(file_name, sheet_name="Insured AR")

    file_AR_RN = file_AR.dropna()
    cell_obj_11 = file_AR_RN(row=1, column=1)
    cell_obj_12 = file_AR_RN(row=1, column=2)
    cell_obj_13 = file_AR_RN(row=1, column=3)
    cell_obj_14 = file_AR_RN(row=1, column=4)
    cell_obj_15 = file_AR_RN(row=1, column=5)
    cell_obj_16 = file_AR_RN(row=1, column=6)
    cell_obj_17 = file_AR_RN(row=1, column=7)
    cell_obj_18 = file_AR_RN(row=1, column=8)

        # below section needs to be expanded with more hard coded values as we see more specific vlaues from Banks
    customer_name = ["Customer", "Customer Name", "Code", "Name", "Client", "Account"]
    upto_30Days = ["30 Days","1-30 past due","Current","30","Current - 30 days","1 - 30","0-30","1-30"]
    upto_60Days = ["60 Days","31-60 past due","30-60 past due","31 - 60 past due","30 - 60 past due","60","30 - 60 days","30 - 60","30-60","31-60","31 - 60","Period 1"]
    upto_90Days = ["90 Days","61-90 past due","60-90 past due","61 - 90 past due","60 - 90 past due","90","60 - 90 days","60 - 90","60-90","61-90","61 - 90","Period 2"]
    above_90Days = ["120 days",">90 past due","91 - 120 days","Over 120 days","Over 90 days","90","Over 90",">90","> 90","90+","Period 3","Period 4"]
    Total_AR = ["Balance","Total","Balance Due","Due"]
        
    if cell_obj_12 in above_90Days:


    if file_AR_RN(row=max_row, column=2) = file_AR_RN.iloc[0:max_row-1].sum() # this has to be modified to sum the rows only from the column
            file_AR = file_AR[:-1]

            Insured_AR_above90Days = file_AR["cell_obj_12"].sum()
    elif cell_obj_13 in above_90Days:

    if file_AR_RN(row=max_row, column=3) = file_AR_RN.iloc[0:max_row-1].sum() # this has to be modified to sum the rows only from the column
            file_AR = file_AR[:-1]

    Insured_AR_above90Days = file_AR["cell_obj_13"].sum() 
    elif cell_obj_14 in above_90Days:
    
    if file_AR_RN(row=max_row, column=4) = file_AR_RN.iloc[0:max_row-1].sum() # this has to be modified to sum the rows only from the column
            file_AR = file_AR[:-1]

    Insured_AR_above90Days = file_AR["cell_obj_14"].sum() 
    elif cell_obj_15 in above_90Days:

    if file_AR_RN(row=max_row, column=5) = file_AR_RN.iloc[0:max_row-1].sum() # this has to be modified to sum the rows only from the column
            file_AR = file_AR[:-1]

    Insured_AR_above90Days = file_AR["cell_obj_15"].sum() 
    elif cell_obj_16 in above_90Days:

    if file_AR_RN(row=max_row, column=6) = file_AR_RN.iloc[0:max_row-1].sum() # this has to be modified to sum the rows only from the column
            file_AR = file_AR[:-1]

    Insured_AR_above90Days = file_AR["cell_obj_16"].sum() 
    elif cell_obj_17 in above_90Days:

    if file_AR_RN(row=max_row, column=7) = file_AR_RN.iloc[0:max_row-1].sum() # this has to be modified to sum the rows only from the column
            file_AR = file_AR[:-1]

    Insured_AR_above90Days = file_AR["cell_obj_13"].sum() 
    else cell_obj_18 in above_90Days:

    if file_AR_RN(row=max_row, column=8) = file_AR_RN.iloc[0:max_row-1].sum() # this has to be modified to sum the rows only from the column
            file_AR = file_AR[:-1]

    Insured_AR_above90Days = file_AR["cell_obj_13"].sum() 

    if cell_obj_12 in Total_AR:
            Insured_AR_Total = file_AR["cell_obj_12"].sum()
    elif cell_obj_13 in Total_AR:
            Insured_AR_Total = file_AR["cell_obj_13"].sum() 
    elif cell_obj_14 in Total_AR:
            Insured_AR_Total = file_AR["cell_obj_14"].sum() 
    elif cell_obj_15 in Total_AR:
            Insured_AR_Total = file_AR["cell_obj_15"].sum() 
    elif cell_obj_16 in Total_AR:
            Insured_AR_Total = file_AR["cell_obj_16"].sum() 
    elif cell_obj_17 in Total_AR:
            Insured_AR_Total = file_AR["cell_obj_13"].sum() 
    else cell_obj_18 in Total_AR:
            Insured_AR_Total = file_AR["cell_obj_13"].sum() 

    # Read uninsured AR
    file_uninsured_AR = pandas.read_excel(
        file_name, sheet_name="Uninsured AR")
    uninsured_ar_total = file_uninsured_AR["Balance"].sum()
    uninsured_ar_Greater_Than_90_Days = file_uninsured_AR["> 90 day"].sum()

    print("Insured Account Receivable Total = ", insured_ar_total)
    print("Insured AR > 90 days = ", AR_Greater_Than_90_Days)

    print("Uninsured Account Receivable Total = ", uninsured_ar_total)
    print("Uninsured AR > 90 days = ", uninsured_ar_Greater_Than_90_Days)

    return AccountsReceivable(
        insured_ar=insured_ar_total,
        insured_ar_90_days=AR_Greater_Than_90_Days,
        uninsured_ar=uninsured_ar_total,
        uninsured_ar_90_days=uninsured_ar_Greater_Than_90_Days,
    )


# Any calculation?
AR_Insured = 0
AR_Contra = 0
AR_Related_Party = 0


def read_acounts_payabale(file_name) -> int:
    file_AP = pandas.read_excel(file_name, sheet_name="AP", skiprows=[0])
    AP_30_60_table = file_AP["30-60"].where(file_AP["Name"] == "Government")
    AP_60_90_table = file_AP["60-90"].where(file_AP["Name"] == "Government")
    AP_GreaterThan_90_table = file_AP["> 90 day"].where(file_AP["Name"] == "Government")

    AP_30_60_table.fillna(0)
    AP_60_90_table.fillna(0)
    AP_GreaterThan_90_table.fillna(0)
    Priority_Payables = (
        AP_30_60_table.sum() + AP_60_90_table.sum() + AP_GreaterThan_90_table.sum()
    )
    return Priority_Payables


def calculate_marginable_ar(Ar, Priority_Payables) -> MarginableAR:
    Marginable_AR = (
        Ar.insured_ar
        - Ar.insured_ar_90_days
        - AR_Contra
        - AR_Related_Party
        - Priority_Payables
    )
    Marginable_AR_90_percent = Marginable_AR * 0.9
    print("Priority Payables = ", Priority_Payables)
    print("Marginable AR = ", Marginable_AR)
    print("90% of Marginable AR = ", Marginable_AR_90_percent)
    print("Contra AR = ", AR_Contra)
    print("Related party AR = ", AR_Related_Party)
    return MarginableAR(marginable_AR_90_percent=Marginable_AR_90_percent, priority_payables=Priority_Payables,Marginable_AR=Marginable_AR,
                        AR_Contra=AR_Contra,AR_Related_Party=AR_Related_Party)


def main():
    file_name = "Sample_File_1.xlsx"
    ar = read_accounts_receivable(file_name=file_name)
    Priority_Payables = read_acounts_payabale(file_name=file_name)
    calculate_marginable_ar(Ar=ar, Priority_Payables=Priority_Payables)


# if __name__ == "__main__":
#     main()
